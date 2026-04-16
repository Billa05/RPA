"""
CertifyHub Event Certificate Platform — Backend
Mock verification service + event/registration management
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime
import random
import string
import hashlib
import io
import os
import openpyxl
from flask import send_from_directory

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")

app = Flask(__name__)
CORS(app)

# ── In-memory database ──────────────────────────────────────────────
DB = {
    "vendors": {
        "vendor1": {"id": "vendor1", "name": "Acme Events", "email": "vendor@acme.com", "password": "vendor123"}
    },
    "events": {},
    "registrations": {},
}

FAILURE_REASONS = [
    "Image is blurry or unreadable",
    "Document appears to be expired",
    "Name on document does not match application",
    "Invalid document format",
    "Document is incomplete or cropped",
    "Signature/seal not visible",
    "Poor lighting — document unreadable",
]

CERT_TYPES = {
    "government_id": "Government ID",
    "photo": "Passport Photo",
    "address_proof": "Address Proof",
    "payment_receipt": "Payment Receipt",
    "education_cert": "Education Certificate",
    "employer_letter": "Employer Letter",
    "medical_cert": "Medical Certificate",
    "noc": "No Objection Certificate",
}

def gen_id(prefix="", length=6):
    chars = string.ascii_uppercase + string.digits
    return prefix + "".join(random.choices(chars, k=length))

def cert_hash(reg_id, name):
    raw = f"{reg_id}:{name}:{datetime.now().isoformat()}"
    return hashlib.sha256(raw.encode()).hexdigest()[:40]


# ══════════════════════════════════════════════════════════
#  VENDOR AUTH
# ══════════════════════════════════════════════════════════

@app.route("/api/vendor/login", methods=["POST"])
def vendor_login():
    data = request.json
    for v in DB["vendors"].values():
        if v["email"] == data.get("email") and v["password"] == data.get("password"):
            return jsonify({"token": v["id"], "vendor": {k: v[k] for k in ("id", "name", "email")}})
    return jsonify({"error": "Invalid credentials"}), 401


# ══════════════════════════════════════════════════════════
#  EVENTS
# ══════════════════════════════════════════════════════════

@app.route("/api/events", methods=["GET"])
def list_events():
    events = list(DB["events"].values())
    # Add registration count to each
    for e in events:
        e["registrationCount"] = sum(
            1 for r in DB["registrations"].values() if r["eventId"] == e["id"]
        )
    return jsonify(events)


@app.route("/api/events/<event_id>", methods=["GET"])
def get_event(event_id):
    event = DB["events"].get(event_id)
    if not event:
        return jsonify({"error": "Event not found"}), 404
    event["registrationCount"] = sum(
        1 for r in DB["registrations"].values() if r["eventId"] == event_id
    )
    return jsonify(event)


@app.route("/api/events", methods=["POST"])
def create_event():
    """
    Multipart form:
      - data (JSON string): name, date, description, venue, requiredDocs[]
      - attendees (Excel file): columns Name, IDNumber
    """
    vendor_id = request.headers.get("X-Vendor-Id")
    if not vendor_id or vendor_id not in DB["vendors"]:
        return jsonify({"error": "Unauthorized"}), 401

    import json as json_mod
    raw = request.form.get("data", "{}")
    data = json_mod.loads(raw)

    # Parse attendee list — either from uploaded Excel OR from typed-in JSON rows
    excel_file = request.files.get("attendees")
    attendees = {}  # { id_number: name }

    if excel_file:
        # Excel upload path
        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            name_col = next(i for i, h in enumerate(headers) if "name" in h)
            id_col   = next(i for i, h in enumerate(headers) if "id" in h)
        except StopIteration:
            return jsonify({"error": "Excel must have 'Name' and 'ID' columns"}), 400

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_col] and row[id_col]:
                attendees[str(row[id_col]).strip()] = str(row[name_col]).strip()
    else:
        # Typed-in rows path — data["attendees"] is a list of {name, idNumber}
        for row in data.get("attendees", []):
            id_no = str(row.get("idNumber", "")).strip()
            name  = str(row.get("name", "")).strip()
            if id_no and name:
                attendees[id_no] = name

    event_id = gen_id("EVT-")
    event = {
        "id": event_id,
        "vendorId": vendor_id,
        "vendorName": DB["vendors"][vendor_id]["name"],
        "name": data.get("name", "Unnamed Event"),
        "date": data.get("date", ""),
        "description": data.get("description", ""),
        "venue": data.get("venue", ""),
        "requiredDocs": data.get("requiredDocs", []),
        "attendees": attendees,
        "attendeeCount": len(attendees),
        "createdAt": datetime.now().isoformat(),
    }
    DB["events"][event_id] = event
    return jsonify({"id": event_id, "message": "Event created", "attendeeCount": len(attendees)}), 201


@app.route("/api/vendor/events", methods=["GET"])
def vendor_events():
    vendor_id = request.headers.get("X-Vendor-Id")
    if not vendor_id:
        return jsonify({"error": "Unauthorized"}), 401
    events = [e for e in DB["events"].values() if e["vendorId"] == vendor_id]
    for e in events:
        e["registrationCount"] = sum(
            1 for r in DB["registrations"].values() if r["eventId"] == e["id"]
        )
    return jsonify(events)


@app.route("/api/vendor/events/<event_id>/registrations", methods=["GET"])
def vendor_event_registrations(event_id):
    vendor_id = request.headers.get("X-Vendor-Id")
    event = DB["events"].get(event_id)
    if not event or event["vendorId"] != vendor_id:
        return jsonify({"error": "Unauthorized or event not found"}), 403
    regs = [r for r in DB["registrations"].values() if r["eventId"] == event_id]
    return jsonify(regs)


# ══════════════════════════════════════════════════════════
#  ATTENDANCE CHECK
# ══════════════════════════════════════════════════════════

@app.route("/api/events/<event_id>/check-attendee", methods=["POST"])
def check_attendee(event_id):
    """
    User provides name + idNumber.
    Returns whether they're in the vendor's Excel attendance list.
    """
    event = DB["events"].get(event_id)
    if not event:
        return jsonify({"error": "Event not found"}), 404

    data = request.json
    id_number = str(data.get("idNumber", "")).strip()
    name      = str(data.get("name", "")).strip().lower()

    attendees = event.get("attendees", {})
    if id_number in attendees:
        stored_name = attendees[id_number].lower()
        if name and name != stored_name:
            return jsonify({
                "found": False,
                "reason": "ID found but name does not match our records"
            })
        # Check if this person already has a registration for this event
        existing_reg = None
        for r in DB["registrations"].values():
            if r["eventId"] == event_id and r["idNumber"] == id_number:
                existing_reg = r
                break

        result = {
            "found": True,
            "attendeeName": attendees[id_number],
            "message": "Attendance confirmed — you may proceed"
        }
        if existing_reg:
            result["existingRegistration"] = existing_reg
        return jsonify(result)

    return jsonify({
        "found": False,
        "reason": "You are not in the attendee list for this event"
    })


# ══════════════════════════════════════════════════════════
#  REGISTRATIONS
# ══════════════════════════════════════════════════════════

@app.route("/api/events/<event_id>/register", methods=["POST"])
def register(event_id):
    """
    Multipart form: field 'data' (JSON) + one file per doc key.
    Falls back to JSON-only for UiPath bot calls.
    """
    event = DB["events"].get(event_id)
    if not event:
        return jsonify({"error": "Event not found"}), 404

    import json as json_mod
    if request.content_type and 'multipart' in request.content_type:
        data = json_mod.loads(request.form.get("data", "{}"))
    else:
        data = request.json or {}

    id_number = str(data.get("idNumber", "")).strip()

    # Re-verify attendance
    if id_number not in event.get("attendees", {}):
        return jsonify({"error": "Not in attendee list"}), 403

    reg_id = gen_id("REG-")

    # Save uploaded files
    reg_dir = os.path.join(UPLOAD_DIR, reg_id)
    os.makedirs(reg_dir, exist_ok=True)

    docs = []
    for d in data.get("documents", []):
        key = d["key"]
        saved_filename = None
        file_obj = request.files.get(key)
        if file_obj and file_obj.filename:
            safe_name = key + "_" + file_obj.filename.replace("/", "_").replace("\\", "_")
            file_obj.save(os.path.join(reg_dir, safe_name))
            saved_filename = safe_name

        docs.append({
            "key": key,
            "name": d.get("name", key),
            "docType": d.get("docType", ""),
            "status": "pending",
            "reason": None,
            "confidence": None,
            "savedFile": saved_filename,
        })

    reg = {
        "id": reg_id,
        "eventId": event_id,
        "eventName": event["name"],
        "applicantName": data.get("name", ""),
        "idNumber": id_number,
        "documents": docs,
        "verificationStatus": "pending",   # pending | partial | failed | verified | approved
        "vendorOverride": False,
        "overrideNote": "",
        "certificateReady": False,
        "certificateHash": None,
        "submittedAt": datetime.now().isoformat(),
    }
    DB["registrations"][reg_id] = reg
    return jsonify({"id": reg_id, "message": "Registration submitted"}), 201


@app.route("/api/registrations/<reg_id>/documents/<filename>", methods=["GET"])
def serve_document(reg_id, filename):
    """Serve an uploaded document file for preview."""
    reg_dir = os.path.join(UPLOAD_DIR, reg_id)
    if not os.path.isdir(reg_dir):
        return jsonify({"error": "No files found"}), 404
    return send_from_directory(reg_dir, filename)


@app.route("/api/registrations/<reg_id>", methods=["GET"])
def get_registration(reg_id):
    reg = DB["registrations"].get(reg_id)
    if not reg:
        return jsonify({"error": "Registration not found"}), 404
    return jsonify(reg)


# ══════════════════════════════════════════════════════════
#  MOCK DOCUMENT VERIFICATION
# ══════════════════════════════════════════════════════════

@app.route("/api/registrations/<reg_id>/verify", methods=["POST"])
def verify_documents(reg_id):
    """
    Mock verification — called after user uploads documents.
    Each document gets randomly validated (70% pass rate).
    Returns per-document result.
    """
    reg = DB["registrations"].get(reg_id)
    if not reg:
        return jsonify({"error": "Registration not found"}), 404

    results = []
    all_passed = True

    for doc in reg["documents"]:
        passed     = random.random() < 0.70
        confidence = round(random.uniform(0.78, 0.99) if passed else random.uniform(0.15, 0.55), 2)
        reason     = None if passed else random.choice(FAILURE_REASONS)

        doc["status"]     = "verified" if passed else "failed"
        doc["confidence"] = confidence
        doc["reason"]     = reason

        if not passed:
            all_passed = False

        results.append({
            "key":        doc["key"],
            "name":       doc["name"],
            "status":     doc["status"],
            "confidence": confidence,
            "reason":     reason,
        })

    if all_passed:
        reg["verificationStatus"] = "verified"
        reg["certificateReady"]   = True
        reg["certificateHash"]    = cert_hash(reg_id, reg["applicantName"])
    else:
        failed_count = sum(1 for d in reg["documents"] if d["status"] == "failed")
        reg["verificationStatus"] = "partial" if failed_count < len(reg["documents"]) else "failed"
        reg["certificateReady"]   = False

    return jsonify({
        "registrationId":     reg_id,
        "results":            results,
        "verificationStatus": reg["verificationStatus"],
        "certificateReady":   reg["certificateReady"],
        "passCount":          sum(1 for r in results if r["status"] == "verified"),
        "failCount":          sum(1 for r in results if r["status"] == "failed"),
    })


# ══════════════════════════════════════════════════════════
#  VENDOR OVERRIDE
# ══════════════════════════════════════════════════════════

@app.route("/api/registrations/<reg_id>/override", methods=["POST"])
def vendor_override(reg_id):
    """
    Vendor can accept a registration even if mock service failed some docs.
    """
    vendor_id = request.headers.get("X-Vendor-Id")
    reg = DB["registrations"].get(reg_id)
    if not reg:
        return jsonify({"error": "Registration not found"}), 404

    # Check vendor owns this event
    event = DB["events"].get(reg["eventId"])
    if not event or event["vendorId"] != vendor_id:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    action = data.get("action", "approve")  # approve | reject

    if action == "approve":
        reg["vendorOverride"]     = True
        reg["overrideNote"]       = data.get("note", "Approved by vendor")
        reg["verificationStatus"] = "approved"
        reg["certificateReady"]   = True
        if not reg["certificateHash"]:
            reg["certificateHash"] = cert_hash(reg_id, reg["applicantName"])
    else:
        reg["vendorOverride"]     = True
        reg["overrideNote"]       = data.get("note", "Rejected by vendor")
        reg["verificationStatus"] = "rejected"
        reg["certificateReady"]   = False

    return jsonify({"message": f"Registration {action}d by vendor", "registration": reg})


# ══════════════════════════════════════════════════════════
#  CERTIFICATE
# ══════════════════════════════════════════════════════════

@app.route("/api/registrations/<reg_id>/certificate", methods=["GET"])
def get_certificate(reg_id):
    reg = DB["registrations"].get(reg_id)
    if not reg:
        return jsonify({"error": "Registration not found"}), 404
    if not reg.get("certificateReady"):
        return jsonify({"error": "Certificate not ready. Complete document verification first."}), 403

    event = DB["events"].get(reg["eventId"], {})
    return jsonify({
        "registrationId":  reg_id,
        "applicantName":   reg["applicantName"],
        "idNumber":        reg["idNumber"],
        "eventName":       event.get("name", ""),
        "eventDate":       event.get("date", ""),
        "eventVenue":      event.get("venue", ""),
        "vendorName":      event.get("vendorName", ""),
        "issuedAt":        datetime.now().isoformat(),
        "hash":            reg["certificateHash"],
        "vendorOverride":  reg["vendorOverride"],
    })


# ══════════════════════════════════════════════════════════
#  SEED DEMO DATA
# ══════════════════════════════════════════════════════════

@app.route("/api/seed", methods=["POST"])
def seed():
    """Creates two demo events with sample attendee lists — no Excel needed."""
    attendees_1 = {
        "ID001": "Rajesh Kumar",
        "ID002": "Priya Sharma",
        "ID003": "Amit Patel",
        "ID004": "Sneha Reddy",
        "ID005": "Mohammed Khan",
    }
    attendees_2 = {
        "TKT101": "Arjun Mehta",
        "TKT102": "Divya Nair",
        "TKT103": "Ravi Shankar",
        "TKT104": "Ananya Bose",
    }

    evt1 = {
        "id": "EVT-DEMO1",
        "vendorId": "vendor1",
        "vendorName": "Acme Events",
        "name": "Tech Summit 2026",
        "date": "2026-04-15",
        "description": "Annual technology conference bringing together industry leaders, developers, and innovators.",
        "venue": "Bangalore International Convention Centre",
        "requiredDocs": ["government_id", "photo", "payment_receipt"],
        "attendees": attendees_1,
        "attendeeCount": len(attendees_1),
        "createdAt": datetime.now().isoformat(),
    }
    evt2 = {
        "id": "EVT-DEMO2",
        "vendorId": "vendor1",
        "vendorName": "Acme Events",
        "name": "National Music Workshop",
        "date": "2026-05-20",
        "description": "A 3-day intensive workshop for classical and contemporary musicians.",
        "venue": "Mumbai Cultural Centre",
        "requiredDocs": ["government_id", "photo", "education_cert"],
        "attendees": attendees_2,
        "attendeeCount": len(attendees_2),
        "createdAt": datetime.now().isoformat(),
    }

    DB["events"]["EVT-DEMO1"] = evt1
    DB["events"]["EVT-DEMO2"] = evt2

    return jsonify({
        "message": "Demo data seeded",
        "events": ["EVT-DEMO1", "EVT-DEMO2"],
        "tip": "Use IDs: ID001-ID005 for Tech Summit, TKT101-TKT104 for Music Workshop"
    }), 201


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "events": len(DB["events"]), "registrations": len(DB["registrations"])})


if __name__ == "__main__":
    print("""
  ╔═══════════════════════════════════════════════╗
  ║   CertifyHub Backend  →  http://localhost:5000 ║
  ╠═══════════════════════════════════════════════╣
  ║  POST  /api/vendor/login                      ║
  ║  GET   /api/events                            ║
  ║  POST  /api/events                (+ Excel)   ║
  ║  POST  /api/events/:id/check-attendee         ║
  ║  POST  /api/events/:id/register               ║
  ║  POST  /api/registrations/:id/verify          ║
  ║  POST  /api/registrations/:id/override        ║
  ║  GET   /api/registrations/:id/certificate     ║
  ║  POST  /api/seed              (demo data)     ║
  ╚═══════════════════════════════════════════════╝
    """)
    app.run(debug=True, host="0.0.0.0", port=5000)
